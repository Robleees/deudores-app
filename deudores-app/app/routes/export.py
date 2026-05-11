import io
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from flask import Blueprint, send_file, session

from app import db
from app.models import Casa, Circuito, Transaccion, Usuario
from app.routes.auth import login_required

export_bp = Blueprint('export', __name__, url_prefix='/export')

_ROJO   = 'FFdc3545'
_VERDE  = 'FF198754'
_GRIS   = 'FFF8F9FA'
_BLANCO = 'FFFFFFFF'
_OSCURO = 'FF1A1A2E'


def _get_usuario_actual():
    return db.session.get(Usuario, session['usuario_id'])


def _header(ws, cols, fill_hex=_OSCURO):
    fill = PatternFill('solid', fgColor=fill_hex)
    font = Font(bold=True, color='FFFFFFFF' if fill_hex == _OSCURO else 'FF000000')
    for col, titulo in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def _ajustar_columnas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


@export_bp.route('/transacciones')
@login_required
def transacciones():
    _get_usuario_actual()

    circuitos = db.session.execute(
        db.select(Circuito).order_by(Circuito.nombre)
    ).scalars().all()

    casas = db.session.execute(
        db.select(Casa).filter_by(activa=True).order_by(Casa.nombre_familia)
    ).scalars().all()

    todas = db.session.execute(
        db.select(Transaccion).order_by(Transaccion.fecha.desc())
    ).scalars().all()

    saldos = {casa.id: casa.saldo_actual() for casa in casas}

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen por circuito ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumen'
    _header(ws1, ['Circuito', 'Descripción', 'Casas activas', 'Total adeudado'])

    for fila, c in enumerate(circuitos, start=2):
        casas_c = [ca for ca in casas if ca.circuito_id == c.id]
        adeudado = sum(saldos[ca.id] for ca in casas_c if saldos[ca.id] > 0)
        ws1.append([c.nombre, c.descripcion or '', len(casas_c), float(adeudado)])
        celda = ws1.cell(row=fila, column=4)
        celda.number_format = '"$"#,##0.00'
        celda.font = Font(color=_ROJO if adeudado > 0 else _VERDE, bold=True)

    _ajustar_columnas(ws1)

    # ── Hoja 2: Saldo por casa ────────────────────────────────────────────
    ws2 = wb.create_sheet('Saldo por casa')
    _header(ws2, ['Circuito', 'Familia', 'Dirección', 'Personas', 'Saldo actual'])

    for fila, ca in enumerate(sorted(casas, key=lambda x: (x.circuito.nombre, x.nombre_familia)), start=2):
        saldo = saldos[ca.id]
        ws2.append([ca.circuito.nombre, ca.nombre_familia, ca.direccion, ca.num_personas, float(saldo)])
        celda = ws2.cell(row=fila, column=5)
        celda.number_format = '"$"#,##0.00'
        celda.font = Font(color=_ROJO if saldo > 0 else _VERDE)
        if fila % 2 == 0:
            for col in range(1, 6):
                ws2.cell(row=fila, column=col).fill = PatternFill('solid', fgColor=_GRIS)

    _ajustar_columnas(ws2)

    # ── Hoja 3: Todas las transacciones ───────────────────────────────────
    ws3 = wb.create_sheet('Transacciones')
    _header(ws3, ['Fecha', 'Circuito', 'Familia', 'Tipo', 'Monto', 'Descripción', 'Registró'])

    rojo_fill  = PatternFill('solid', fgColor='FFFCE8E8')
    verde_fill = PatternFill('solid', fgColor='FFE8F5E9')

    casa_index = {ca.id: ca for ca in casas}

    for fila, t in enumerate(todas, start=2):
        ca = casa_index.get(t.casa_id)
        circuito_nombre = ca.circuito.nombre if ca else '—'
        familia = ca.nombre_familia if ca else '—'
        ws3.append([
            t.fecha.strftime('%d/%m/%Y %H:%M'),
            circuito_nombre,
            familia,
            'Cargo' if t.tipo == 'cargo' else 'Abono',
            float(t.monto),
            t.descripcion or '',
            t.usuario.nombre,
        ])
        celda_monto = ws3.cell(row=fila, column=5)
        celda_monto.number_format = '"$"#,##0.00'
        fondo = rojo_fill if t.tipo == 'cargo' else verde_fill
        for col in range(1, 8):
            ws3.cell(row=fila, column=col).fill = fondo

    _ajustar_columnas(ws3)

    # ── Enviar archivo ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"deudas_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
